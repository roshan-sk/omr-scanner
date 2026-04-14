import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import OMRSheet, OMRAnswer


def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_len:
                    max_len = len(value)
            except:
                pass

        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)


def build_excel(sheet_ids=None, latest_only=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    sheets_query = OMRSheet.query

    if sheet_ids:
        sheets_query = sheets_query.filter(OMRSheet.id.in_(sheet_ids))

    sheets = sheets_query.order_by(OMRSheet.id.desc()).all()

    if not sheets:
        return None

    first_answers = (
        OMRAnswer.query
        .filter_by(sheet_id=sheets[0].id)
        .order_by(OMRAnswer.question_no)
        .all()
    )
    total_questions = len(first_answers)

    headers = ["Roll Number", "File Name"]

    for i in range(1, total_questions + 1):
        headers.append(f"Q{str(i).zfill(3)}")

    headers += ["Correct", "Wrong", "Percentage"]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for sheet in sheets:
        row = [
            sheet.roll_number,
            sheet.result_file
        ]

        answers = (
            OMRAnswer.query
            .filter_by(sheet_id=sheet.id)
            .order_by(OMRAnswer.question_no)
            .all()
        )

        for ans in answers:
            row.append(ans.selected_option)

        row += [
            sheet.correct_answers,
            sheet.wrong_answers,
            round(sheet.percentage, 2)
        ]

        ws.append(row)

    auto_fit_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f"omr_results_file_{timestamp}.xlsx"

    return output, filename